import openpyxl
import os
import random

desktop_location = "C:/Users/jyuill/Desktop/"
file_name = "jasonwb.xlsx"
final_name = os.path.join(desktop_location, file_name)

wb = openpyxl.workbook.Workbook()  # this will make a workbook AND a worksheet
wb.remove(wb.active)

tabcolors = ["ff80ed", "065535", "133337", "ffffff"]
worksheetnames = ["first", "second", "third", "fourth"]

if len(tabcolors) != len(worksheetnames):
    raise Exception("lens do not match")

counter = 0
for name in worksheetnames:
    wb.create_sheet(name)
    wb[name].sheet_properties.tabColor = tabcolors[counter]
    counter += 1








wb.save(final_name)
