from openpyxl import Workbook, load_workbook

wb = load_workbook('test.xlsx')

# get the sheets
ws = wb.active
#print(ws['A1'].value)


# get different sheets
print(wb['ts3'])

wb.create_sheet("testasfsf")
print(wb.sheetnames)

wb.save('test.xlsx')