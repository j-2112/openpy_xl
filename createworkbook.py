# lets create a workbook in python! yaya

# make sure we have open py xl installed, and then import the workbook from it

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# next, lets make a wrokbook


wb = load_workbook('createwb.xlsx')
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)