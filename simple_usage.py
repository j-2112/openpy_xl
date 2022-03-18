# Simiple useage tutorial from read the docs

# write a workbook
import openpyxl

wb = openpyxl.Workbook()
dest_filename = 'empty_book.xlsx'

ws1=wb.active

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title= "Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="thrid_worksheet_data")

wb.save(filename= dest_filename)
